"""LKE writer — isi Lembar Kerja Evaluasi (.xlsx) TANPA merusak rumus.

Dipakai skill evaluasi yang berbasis LKE ber-formula (SAKIP, SPIP). Prinsip:
hanya menulis CELL INPUT; setiap percobaan menulis ke CELL FORMULA atau sheet
agregator DITOLAK (bukan ditimpa). Rumus & sheet agregator tidak pernah berubah.

Deteksi formula berlapis:
  1. cell-map (`cell-map-formulas.json` → formula_cells_all per sheet) bila ada.
  2. runtime: `cell.data_type == "f"` (universal, tak perlu cell-map).

Port v7-native dari knowledge/skills/evaluasi-spip/references/fill_lke_safely.py
(dibuat generik untuk LKE apa pun, cell-map opsional).
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


class LKEWriter:
    def __init__(self, path: str | Path, cellmap_path: str | Path | None = None,
                 aggregator_sheets: set[str] | None = None):
        self.path = Path(path)
        # data_only=False → formula DIPERTAHANKAN sebagai string (bukan nilai cache).
        self.wb = load_workbook(self.path, data_only=False, keep_vba=False)
        self.formula_map: dict[str, set[str]] = {}
        if cellmap_path and Path(cellmap_path).is_file():
            try:
                data = json.loads(Path(cellmap_path).read_text(encoding="utf-8"))
                for sn, d in data.items():
                    self.formula_map[sn] = set(d.get("formula_cells_all", []))
            except (json.JSONDecodeError, OSError):
                self.formula_map = {}
        self.aggregators = set(aggregator_sheets or [])
        self.written: list[dict] = []
        self.refused: list[dict] = []

    def set(self, sheet: str, coord: str, value: Any, note: str | None = None) -> bool:
        """Tulis value ke sheet!coord bila AMAN (bukan formula/agregator). Return True bila ditulis."""
        coord = str(coord).strip().upper()
        if sheet not in self.wb.sheetnames:
            self.refused.append({"sheet": sheet, "coord": coord, "alasan": "sheet tidak ada"})
            return False
        if sheet in self.aggregators:
            self.refused.append({"sheet": sheet, "coord": coord, "alasan": "sheet agregator (formula-only)"})
            return False
        if coord in self.formula_map.get(sheet, set()):
            self.refused.append({"sheet": sheet, "coord": coord, "alasan": "cell formula (cell-map)"})
            return False
        ws = self.wb[sheet]
        try:
            cell = ws[coord]
        except (ValueError, KeyError):
            self.refused.append({"sheet": sheet, "coord": coord, "alasan": "koordinat invalid"})
            return False
        if cell.data_type == "f":
            self.refused.append({"sheet": sheet, "coord": coord, "alasan": "cell formula (runtime)"})
            return False
        cell.value = value
        self.written.append({"sheet": sheet, "coord": coord, "value": value, "note": note or ""})
        return True

    def save(self, out_path: str | Path) -> Path:
        out = Path(out_path)
        out.parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(out)
        return out
