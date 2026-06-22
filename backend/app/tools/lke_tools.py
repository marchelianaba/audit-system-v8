"""Tool fill_lke — isi Lembar Kerja Evaluasi (LKE) Excel untuk skill evaluasi
(SAKIP/SPIP) TANPA mengubah rumus. Hanya cell input yang ditulis; cell formula
& sheet agregator DITOLAK.

Sumber LKE:
  - SPIP: template ber-rumus yang dibawa skill (knowledge/skills/evaluasi-spip/
    references/templates/lke-spip-kementerian.xlsx + cell-map).
  - SAKIP / lainnya: LKE .xlsx yang DIUPLOAD auditor ke folder penugasan.
Output ditulis ke salinan kerja `_KKP/LKE-terisi-<skill>.xlsx` (template/upload
asli tidak pernah diubah).
"""
import json
import re
import shutil
from pathlib import Path

from claude_agent_sdk import tool
from openpyxl import load_workbook

from app.config import get_settings
from app.lke_writer import LKEWriter

# Sheet agregator SPIP (formula-only) — tak boleh ditulis.
_SPIP_AGGREGATORS = {"KKlead I KL", "KKLEAD II", "KKLEAD III", "KKLEAD_SPIP"}


def _slug(skill: str) -> str:
    return re.sub(r"[^a-z0-9\-]", "-", str(skill).strip().lower())


def _spip_template() -> tuple[Path, Path]:
    base = get_settings().skills_path / "evaluasi-spip" / "references" / "templates"
    return base / "lke-spip-kementerian.xlsx", base / "cell-map-formulas.json"


def _find_uploaded_xlsx(folder: Path) -> Path | None:
    """Cari LKE .xlsx yang diupload auditor di subfolder input penugasan.

    Abaikan output kerja (_KKP/_LHP/_QA-SAIPI) supaya tidak memungut hasil sendiri.
    """
    skip = {"_KKP", "_LHP", "_QA-SAIPI", "_INGESTED"}
    candidates: list[Path] = []
    for p in folder.rglob("*.xlsx"):
        if any(part in skip for part in p.relative_to(folder).parts):
            continue
        if p.name.startswith("~$"):  # lock file Excel
            continue
        candidates.append(p)
    return sorted(candidates)[0] if candidates else None


def _resolve_source(folder: Path, skill: str) -> tuple[Path | None, Path | None, str]:
    """Tentukan (source_xlsx, cellmap, note). Prefer upload auditor; SPIP fallback ke template."""
    uploaded = _find_uploaded_xlsx(folder)
    if uploaded is not None:
        return uploaded, None, f"LKE dari upload: {uploaded.name}"
    if _slug(skill) == "evaluasi-spip":
        tpl, cmap = _spip_template()
        if tpl.is_file():
            return tpl, (cmap if cmap.is_file() else None), f"LKE template SPIP: {tpl.name}"
    return None, None, "tidak ada LKE — upload file .xlsx LKE dulu"


@tool(
    "fill_lke",
    "Isi Lembar Kerja Evaluasi (LKE) Excel untuk skill evaluasi (SAKIP/SPIP) TANPA "
    "mengubah rumus — hanya cell INPUT yang ditulis; cell formula & sheet agregator "
    "DITOLAK otomatis (dilaporkan di 'refused'). Sumber: LKE .xlsx yang diupload "
    "auditor, atau template SPIP bawaan. Output: _KKP/LKE-terisi-<skill>.xlsx (asli "
    "tak diubah). `entries` = list of {sheet, coord (mis. 'K6'), value, note?}. "
    "Pakai SEBELUM menyusun catatan/temuan untuk skill ber-LKE.",
    {"penugasan_folder": str, "skill": str, "entries": list},
)
async def fill_lke(args: dict) -> dict:
    folder = Path(args["penugasan_folder"])
    skill = str(args.get("skill", "")).strip()
    entries = args.get("entries") or []
    if not isinstance(entries, list) or not entries:
        return {"content": [{"type": "text", "text": "FAILED|entries kosong (list of {sheet,coord,value})"}], "is_error": True}

    src, cellmap, note = _resolve_source(folder, skill)
    if src is None:
        return {"content": [{"type": "text", "text": f"FAILED|{note}"}], "is_error": True}

    out = folder / "_KKP" / f"LKE-terisi-{_slug(skill)}.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    # Mulai dari salinan kerja yang sudah ada (akumulatif) bila ada, else dari source.
    base = out if out.is_file() else src
    try:
        writer = LKEWriter(
            base, cellmap_path=cellmap,
            aggregator_sheets=_SPIP_AGGREGATORS if _slug(skill) == "evaluasi-spip" else None,
        )
    except Exception as e:  # noqa: BLE001
        return {"content": [{"type": "text", "text": f"FAILED|gagal buka LKE: {e}"}], "is_error": True}

    for e in entries:
        if not isinstance(e, dict):
            continue
        writer.set(str(e.get("sheet", "")), str(e.get("coord", "")), e.get("value"), e.get("note"))
    writer.save(out)

    payload = {
        "ok": True,
        "sumber": note,
        "output": str(out.relative_to(folder)),
        "ditulis": len(writer.written),
        "ditolak_count": len(writer.refused),
        "ditolak_formula": writer.refused[:40],  # cell ditolak (formula/agregator) — pilih cell input lain
    }
    return {"content": [{"type": "text", "text": json.dumps(payload, ensure_ascii=False)}]}


def _lke_to_read(folder: Path, skill: str) -> tuple[Path | None, str]:
    """LKE yang dibaca: salinan kerja _KKP/LKE-terisi (bila ada, paling mutakhir),
    else sumber (upload auditee / template)."""
    work = folder / "_KKP" / f"LKE-terisi-{_slug(skill)}.xlsx"
    if work.is_file():
        return work, f"salinan kerja: {work.name}"
    src, _cmap, note = _resolve_source(folder, skill)
    return src, note


@tool(
    "read_lke",
    "Baca isi Lembar Kerja Evaluasi (LKE) auditee untuk skill evaluasi (SAKIP/SPIP) — "
    "self-assessment yang sudah diisi auditee. Tanpa `sheet`: daftar nama sheet + "
    "jumlah cell terisi. Dengan `sheet`: nilai cell non-kosong (coord→{v,f}; f=true bila "
    "FORMULA, jangan ditulis). Pakai untuk MENILAI self-assessment auditee sebelum "
    "mengisi kolom APIP via fill_lke.",
    {"penugasan_folder": str, "skill": str, "sheet": str},
)
async def read_lke(args: dict) -> dict:
    folder = Path(args["penugasan_folder"])
    skill = str(args.get("skill", "")).strip()
    sheet = str(args.get("sheet", "")).strip()
    src, note = _lke_to_read(folder, skill)
    if src is None or not Path(src).is_file():
        return {"content": [{"type": "text", "text": f"FAILED|{note}"}], "is_error": True}
    try:
        wb = load_workbook(src, data_only=False, read_only=False)
    except Exception as e:  # noqa: BLE001
        return {"content": [{"type": "text", "text": f"FAILED|gagal buka LKE: {e}"}], "is_error": True}

    if not sheet:
        sheets = []
        for sn in wb.sheetnames:
            ws = wb[sn]
            filled = sum(1 for row in ws.iter_rows() for c in row if c.value not in (None, ""))
            sheets.append({"sheet": sn, "terisi": filled, "dim": ws.dimensions})
        payload = {"sumber": note, "total_sheet": len(sheets), "sheets": sheets}
        return {"content": [{"type": "text", "text": json.dumps(payload, ensure_ascii=False)}]}

    if sheet not in wb.sheetnames:
        return {"content": [{"type": "text", "text": f"NOT_FOUND|sheet '{sheet}'. Ada: {wb.sheetnames}"}], "is_error": True}
    ws = wb[sheet]
    # Bound DATA (bukan slice string JSON): max 150 cell, nilai dipotong 60 char.
    _CAP = 150
    cells: dict[str, dict] = {}
    capped = False
    for row in ws.iter_rows():
        for c in row:
            if c.value in (None, ""):
                continue
            if len(cells) >= _CAP:
                capped = True
                break
            cells[c.coordinate] = {"v": str(c.value)[:60], "f": c.data_type == "f"}
        if capped:
            break
    payload = {"sumber": note, "sheet": sheet, "n_cell": len(cells), "capped": capped, "cells": cells}
    return {"content": [{"type": "text", "text": json.dumps(payload, ensure_ascii=False)}]}


@tool(
    "write_penilaian_lke",
    "Tulis (overwrite) _KKP/penilaian-lke-<skill>.json — REKAP skor/predikat hasil "
    "evaluasi ber-LKE (SAKIP/SPIP), sumber tunggal untuk rekap di KKP. Panggil SEKALI "
    "setelah SEMUA unsur/komponen dinilai & fill_lke selesai. Struktur `penilaian` = "
    "{komponen:[{nama, bobot, nilai_pm, nilai_apip, predikat, catatan?}], total_pm?, "
    "total_apip?, predikat_akhir?}. nilai_pm = penilaian mandiri auditee, nilai_apip = "
    "hasil penjaminan APIP. render_kkp_docx akan menampilkan tabel 'Rekap Penilaian (LKE)' "
    "dari file ini di atas daftar AoI.",
    {"penugasan_folder": str, "skill": str, "penilaian": dict},
)
async def write_penilaian_lke(args: dict) -> dict:
    folder = Path(args["penugasan_folder"])
    skill = _slug(args.get("skill", ""))
    pen = args.get("penilaian") or {}
    if not skill:
        return {"content": [{"type": "text", "text": "FAILED|skill kosong"}], "is_error": True}
    out_dir = folder / "_KKP"
    out_dir.mkdir(parents=True, exist_ok=True)
    out = out_dir / f"penilaian-lke-{skill}.json"
    pen.setdefault("skill", skill)
    out.write_text(json.dumps(pen, ensure_ascii=False, indent=2), encoding="utf-8")
    n = len(pen.get("komponen", []) or [])
    return {"content": [{"type": "text", "text": f"OK|penilaian-lke ditulis|n_komponen={n}|{out.relative_to(folder)}"}]}


LKE_TOOLS = [read_lke, fill_lke, write_penilaian_lke]
