"""
fill_lke_safely.py — Helper mengisi LKE SPIP Kementerian tanpa merusak formula.

Kontrak penggunaan:
    from fill_lke_safely import LKEWriter

    w = LKEWriter("LKE SPIP KEMENTERIAN.xlsx", backup=True)
    w.set("KKE 1.1 SASTRA PEMDA", "K6", "Y")
    w.set("KK3.1", "V12", 4.0, note="Direvisi dari modus 3.0 ke 4.0 karena ...")
    w.set("KK4_PENALTI", "C5", "YA")
    w.set("KK4_PENALTI", "D5", 2.0)
    w.save("LKE SPIP KEMENTERIAN - PK.xlsx")

Setiap `set()` akan GAGAL (raise) jika:
  - Sheet bernama agregator (KKlead I KL, KKLEAD II, KKLEAD III, KKLEAD_SPIP, Uraian NIlai Setiap Unsur)
  - Cell target bertipe formula
  - Cell berada di zona yang di-blacklist

Tidak ada delete row/col, tidak ada sheet add/remove.
"""

from __future__ import annotations
import json
import shutil
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


AGGREGATOR_SHEETS = {
    "KKlead I KL",
    "KKLEAD II",
    "KKLEAD III",
    "KKLEAD_SPIP",
    # Uraian Nilai Setiap Unsur: kolom M boleh diisi; sheet tidak sepenuhnya terlarang
    # Jadi hanya diberi peringatan per-cell, bukan blok sheet.
}

# Sheet input yang dikenali — selain ini akan diberi warning
KNOWN_INPUT_SHEETS = {
    "KKE 1.1 SASTRA PEMDA",
    "KKE 1.2 SASARAN OPD",
    "KKE 2.1 SASKEG",
    "KK 2.2 RO",
    "KKE 2.2 KEGIATAN",
    "KK3.1", "KK3.2", "KK3.3", "KK3.4",
    "KK 5.1A", "KK 5.1 B ", "KK 5.2 ",
    "KK 6", "KK 7", "KK 8",
    "KK4_PENALTI",
    "qa 3.1 8 satker",
    "Uraian NIlai Setiap Unsur",
}


class LKEWriter:
    def __init__(self, path: str | Path, backup: bool = True,
                 cellmap_path: Optional[str | Path] = None):
        self.src = Path(path)
        if not self.src.exists():
            raise FileNotFoundError(self.src)

        if backup:
            bak = self.src.with_suffix(self.src.suffix + ".bak")
            if not bak.exists():
                shutil.copy2(self.src, bak)
                print(f"[backup] {bak}")

        self.wb = load_workbook(self.src, data_only=False, keep_vba=False)

        # Muat daftar formula (optional tapi direkomendasikan)
        self.formula_map: dict[str, set[str]] = {}
        if cellmap_path is None:
            # Auto-resolve: cari di folder templates dekat script
            here = Path(__file__).parent
            candidate = here / "templates" / "cell-map-formulas.json"
            if candidate.exists():
                cellmap_path = candidate
        if cellmap_path and Path(cellmap_path).exists():
            data = json.loads(Path(cellmap_path).read_text(encoding="utf-8"))
            for sn, d in data.items():
                self.formula_map[sn] = set(d.get("formula_cells_all", []))
            print(f"[cellmap] loaded {sum(len(v) for v in self.formula_map.values())} "
                  f"formula cells across {len(self.formula_map)} sheets")

        self._log: list[tuple[str, str, Any]] = []

    # -------- internal --------

    def _require_sheet(self, sheet: str) -> Worksheet:
        if sheet not in self.wb.sheetnames:
            raise KeyError(f"Sheet '{sheet}' tidak ditemukan. "
                           f"Tersedia: {self.wb.sheetnames}")
        if sheet in AGGREGATOR_SHEETS:
            raise PermissionError(
                f"Sheet '{sheet}' adalah agregator (formula-only). "
                f"TIDAK BOLEH ditulis. Edit sheet input saja."
            )
        if sheet not in KNOWN_INPUT_SHEETS:
            print(f"[warn] Sheet '{sheet}' tidak dikenali sebagai sheet input. Lanjut.")
        return self.wb[sheet]

    def _assert_writable(self, sheet: str, coord: str) -> None:
        # Cek dari cellmap
        formulas = self.formula_map.get(sheet, set())
        if coord in formulas:
            raise PermissionError(
                f"REFUSE: {sheet}!{coord} adalah FORMULA (per cell-map). Jangan tulis."
            )
        # Cek real-time dari workbook
        ws = self.wb[sheet]
        cell = ws[coord]
        if cell.data_type == "f":
            raise PermissionError(
                f"REFUSE: {sheet}!{coord} bertipe formula saat runtime. Jangan tulis."
            )

    # -------- public --------

    def set(self, sheet: str, coord: str, value: Any,
            note: Optional[str] = None) -> None:
        """Tulis value ke sheet!coord. Gagal kalau target adalah formula."""
        self._require_sheet(sheet)
        self._assert_writable(sheet, coord)
        self.wb[sheet][coord].value = value
        self._log.append((sheet, coord, value))
        if note:
            print(f"[set]  {sheet}!{coord} = {value!r}  # {note}")
        else:
            print(f"[set]  {sheet}!{coord} = {value!r}")

    def set_row(self, sheet: str, row: int, mapping: dict[str, Any]) -> None:
        """Set banyak cell sekaligus: mapping {kolom: value}, misal {'K': 'Y', 'L': 'Y'}."""
        for col, val in mapping.items():
            self.set(sheet, f"{col}{row}", val)

    def save(self, out_path: Optional[str | Path] = None) -> Path:
        out = Path(out_path) if out_path else self.src
        self.wb.save(out)
        print(f"[save] {out}  ({len(self._log)} cell ditulis)")
        return out

    def audit_log(self) -> list[tuple[str, str, Any]]:
        return list(self._log)


# ---- Contoh ----
if __name__ == "__main__":
    import sys
    path = sys.argv[1] if len(sys.argv) > 1 else "LKE SPIP KEMENTERIAN.xlsx"
    w = LKEWriter(path)
    print("\nContoh: tulis Y/T pada PK SASTRA baris 6")
    # w.set_row("KKE 1.1 SASTRA PEMDA", 6, {"K": "Y", "L": "Y", "M": "Y", "N": "Y", "O": "Y"})
    # w.save("LKE SPIP KEMENTERIAN - PK.xlsx")
