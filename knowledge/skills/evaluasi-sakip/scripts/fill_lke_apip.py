"""
fill_lke_apip.py

Membaca LKE SAKIP asli (.xls), mengisi kolom evaluasi APIP berdasarkan
file JSON hasil evaluasi, lalu menyimpan sebagai .xlsx.

Penggunaan:
  python fill_lke_apip.py <lke_asli.xls> <evaluasi.json> <output.xlsx>
"""

import sys
import json
import xlrd
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Helpers: konversi warna xlrd -> openpyxl
# ---------------------------------------------------------------------------

def xls_color_to_hex(xf_index, workbook):
    """Konversi index warna XF dari xlrd ke string hex openpyxl."""
    try:
        xf = workbook.xf_list[xf_index]
        bg = xf.background
        if bg is not None:
            pattern_color_index = bg.pattern_colour_index
            if pattern_color_index not in (0x40, 0x41, None):
                color_map = workbook.colour_map
                rgb = color_map.get(pattern_color_index)
                if rgb is not None:
                    return "{:02X}{:02X}{:02X}".format(*rgb)
    except Exception:
        pass
    return None


def xls_font_color_to_hex(xf_index, workbook):
    """Konversi warna font dari XF xlrd ke string hex."""
    try:
        xf = workbook.xf_list[xf_index]
        font = workbook.font_list[xf.font_index]
        color_index = font.colour_index
        if color_index not in (0x40, 0x41, None, 0x7FFF):
            color_map = workbook.colour_map
            rgb = color_map.get(color_index)
            if rgb is not None:
                return "{:02X}{:02X}{:02X}".format(*rgb)
    except Exception:
        pass
    return None


def border_style_name(style_int):
    """Map integer border style xlrd ke nama openpyxl."""
    mapping = {
        1: "thin",
        2: "medium",
        3: "thick",
        4: "dashed",
        5: "dotted",
        6: "double",
        7: "hair",
        8: "mediumDashed",
        9: "dashDot",
        10: "mediumDashDot",
        11: "dashDotDot",
        12: "mediumDashDotDot",
        13: "slantDashDot",
    }
    return mapping.get(style_int, None)


def make_side(border_type, color_index, workbook):
    style = border_style_name(border_type)
    if style is None:
        return Side(style=None)
    hex_color = None
    try:
        color_map = workbook.colour_map
        rgb = color_map.get(color_index)
        if rgb is not None:
            hex_color = "{:02X}{:02X}{:02X}".format(*rgb)
    except Exception:
        pass
    if hex_color:
        return Side(style=style, color=hex_color)
    return Side(style=style)


# ---------------------------------------------------------------------------
# Salin LKE xls -> xlsx (struktur + format)
# ---------------------------------------------------------------------------

def copy_xls_to_xlsx(xls_path):
    """Baca .xls dengan xlrd, salin ke openpyxl Workbook."""
    wb_xls = xlrd.open_workbook(xls_path, formatting_info=True)
    wb_xlsx = openpyxl.Workbook()

    # Hapus sheet default openpyxl
    for sheet_name in wb_xlsx.sheetnames:
        del wb_xlsx[sheet_name]

    for sheet_idx in range(wb_xls.nsheets):
        ws_xls = wb_xls.sheet_by_index(sheet_idx)
        ws_xlsx = wb_xlsx.create_sheet(title=ws_xls.name)

        # Lebar kolom
        for col_idx in range(ws_xls.ncols):
            col_info = ws_xls.colinfo_map.get(col_idx)
            if col_info is not None:
                # xlrd menyimpan lebar dalam satuan 1/256 karakter
                width_chars = col_info.width / 256.0
                if width_chars > 0:
                    ws_xlsx.column_dimensions[
                        get_column_letter(col_idx + 1)
                    ].width = width_chars

        # Tinggi baris
        for row_idx in range(ws_xls.nrows):
            row_info = ws_xls.rowinfo_map.get(row_idx)
            if row_info is not None and row_info.height_mismatch:
                ws_xlsx.row_dimensions[row_idx + 1].height = (
                    row_info.height / 20.0
                )

        # Merge cells
        for crange in ws_xls.merged_cells:
            rlo, rhi, clo, chi = crange
            if (rhi - rlo > 1) or (chi - clo > 1):
                ws_xlsx.merge_cells(
                    start_row=rlo + 1,
                    start_column=clo + 1,
                    end_row=rhi,
                    end_column=chi,
                )

        # Salin isi dan format sel
        for row_idx in range(ws_xls.nrows):
            for col_idx in range(ws_xls.ncols):
                cell_xls = ws_xls.cell(row_idx, col_idx)
                cell_xlsx = ws_xlsx.cell(row=row_idx + 1, column=col_idx + 1)

                # Nilai
                xtype = cell_xls.ctype
                if xtype == xlrd.XL_CELL_TEXT:
                    cell_xlsx.value = cell_xls.value
                elif xtype == xlrd.XL_CELL_NUMBER:
                    cell_xlsx.value = cell_xls.value
                elif xtype == xlrd.XL_CELL_DATE:
                    import datetime
                    dt = xlrd.xldate_as_datetime(
                        cell_xls.value, wb_xls.datemode
                    )
                    cell_xlsx.value = dt
                elif xtype == xlrd.XL_CELL_BOOLEAN:
                    cell_xlsx.value = bool(cell_xls.value)
                elif xtype == xlrd.XL_CELL_ERROR:
                    cell_xlsx.value = None
                else:
                    try:
                        cell_xlsx.value = cell_xls.value if cell_xls.value else None
                    except AttributeError:
                        pass  # MergedCell — skip

                # Format
                xf_index = cell_xls.xf_index
                if xf_index is None:
                    continue

                try:
                    xf = wb_xls.xf_list[xf_index]
                except IndexError:
                    continue

                # Background fill
                bg_hex = xls_color_to_hex(xf_index, wb_xls)
                if bg_hex:
                    cell_xlsx.fill = PatternFill(
                        fill_type="solid", fgColor=bg_hex
                    )

                # Font
                try:
                    font_xls = wb_xls.font_list[xf.font_index]
                    font_color = xls_font_color_to_hex(xf_index, wb_xls)
                    font_kwargs = {
                        "name": font_xls.name,
                        "size": font_xls.height / 20.0,
                        "bold": bool(font_xls.bold),
                        "italic": bool(font_xls.italic),
                        "underline": (
                            "single" if font_xls.underline_type == 1 else None
                        ),
                    }
                    if font_color:
                        font_kwargs["color"] = font_color
                    cell_xlsx.font = Font(**font_kwargs)
                except Exception:
                    pass

                # Alignment
                try:
                    align = xf.alignment
                    h_map = {
                        0: "general", 1: "left", 2: "center",
                        3: "right", 4: "fill", 5: "justify",
                        6: "centerContinuous", 7: "distributed",
                    }
                    v_map = {
                        0: "top", 1: "center", 2: "bottom",
                        3: "justify", 4: "distributed",
                    }
                    cell_xlsx.alignment = Alignment(
                        horizontal=h_map.get(align.hor_align, "general"),
                        vertical=v_map.get(align.vert_align, "bottom"),
                        wrap_text=bool(align.text_wrapped),
                    )
                except Exception:
                    pass

                # Border
                try:
                    brd = xf.border
                    cell_xlsx.border = Border(
                        left=make_side(
                            brd.left_colour_index,
                            brd.left_colour_index,
                            wb_xls,
                        ),
                        right=make_side(
                            brd.right_colour_index,
                            brd.right_colour_index,
                            wb_xls,
                        ),
                        top=make_side(
                            brd.top_colour_index,
                            brd.top_colour_index,
                            wb_xls,
                        ),
                        bottom=make_side(
                            brd.bottom_colour_index,
                            brd.bottom_colour_index,
                            wb_xls,
                        ),
                    )
                except Exception:
                    pass

    return wb_xlsx, wb_xls


# ---------------------------------------------------------------------------
# Cari baris berdasarkan konten kolom A atau B
# ---------------------------------------------------------------------------

def find_row_by_kode(ws, kode, search_cols=(1, 2)):
    """Cari baris yang mengandung kode tertentu di kolom A atau B."""
    for row in ws.iter_rows():
        for col_idx in search_cols:
            if col_idx - 1 < len(row):
                cell = row[col_idx - 1]
                val = cell.value
                if val is not None and str(val).strip() == str(kode).strip():
                    return cell.row
    return None


def find_row_by_content(ws, content, col=1):
    """Cari baris berdasarkan konten substring di kolom tertentu."""
    for row in ws.iter_rows(min_col=col, max_col=col):
        cell = row[0]
        if cell.value is not None and str(content).lower() in str(
            cell.value
        ).lower():
            return cell.row
    return None


# ---------------------------------------------------------------------------
# Tulis nilai ke sel (kolom offset dari baris dasar)
# ---------------------------------------------------------------------------

def write_cell(ws, row, col, value):
    """Tulis nilai ke sel, pertahankan format yang ada."""
    if row is None:
        return
    cell = ws.cell(row=row, column=col)
    cell.value = value


# ---------------------------------------------------------------------------
# Logika pengisian berdasarkan struktur JSON
# ---------------------------------------------------------------------------

COL_A = 1
COL_B = 2
COL_C = 3

# Mapping kolom relatif untuk tiap level (1-indexed, kolom C = 3)
# Disesuaikan dengan struktur LKE APIP standar Kemenpan-RB
# Row offset dihitung dari baris header kode item

KRITERIA_COL = {
    "predikat_apip": COL_C + 9,    # kolom C+9 = L (asumsi C=3, L=12)
    "nilai_apip": COL_C + 10,      # M = 13
    "keterangan_apip": COL_C + 15, # R = 18
}

# Offset kolom untuk sub-komponen dan komponen disesuaikan LKE APIP Kemenpan
# (kolom diisi berdasarkan posisi header: C12, C13, C14, C16, C17)
# Angka 12, 13, dst. merujuk ke kolom ke-12, 13 dst. di Excel (L, M, N, P, Q, S, T)
SUBKOMP_COL = {
    "persen_terpenuhi": 12,  # L
    "nilai_apip": 13,        # M
    "predikat_apip": 14,     # N
    "nilai_akhir_apip": 16,  # P
    "persen_apip": 17,       # Q
}

KOMP_COL = {
    "nilai_apip": 16,           # P
    "persen_apip": 17,          # Q
    "catatan_evaluator": 19,    # S
    "rekomendasi_evaluator": 20, # T
}

TOTAL_COL = {
    "nilai_total_apip": 16,  # P
    "persen_total_apip": 17, # Q
}


def isi_lke_dari_json(ws, data_json):
    """
    Isi kolom evaluasi APIP di worksheet berdasarkan data JSON.

    Pendekatan: scan baris worksheet dengan context tracking sub-komponen,
    cocokkan kode/nomor dengan data JSON, tulis nilai ke kolom yang sesuai.
    """
    import re

    # Bangun indeks: kode -> data, untuk komponen, sub-komponen, kriteria
    index_komp = {}
    index_subkomp = {}
    # index_kriteria[(kode_sub, nomor)] = krit
    index_kriteria = {}

    for komp in data_json.get("komponen", []):
        index_komp[komp["kode"]] = komp
        for sub in komp.get("sub_komponen", []):
            index_subkomp[sub["kode"]] = sub
            for krit in sub.get("kriteria", []):
                key = (sub["kode"], krit["nomor"])
                index_kriteria[key] = krit

    # Nilai total
    nilai_total = data_json.get("nilai_total", {})

    # Context tracking
    current_subkomp_kode = None

    # Scan baris worksheet
    max_row = ws.max_row
    for row_idx in range(1, max_row + 1):
        cell_a = ws.cell(row=row_idx, column=COL_A)
        cell_b = ws.cell(row=row_idx, column=COL_B)
        cell_c = ws.cell(row=row_idx, column=COL_C)

        val_a = str(cell_a.value).strip() if cell_a.value is not None else ""
        val_b = str(cell_b.value).strip() if cell_b.value is not None else ""
        val_c = str(cell_c.value).strip() if cell_c.value is not None else ""

        # --- Baris total (mengandung "NILAI TOTAL") ---
        if "nilai total" in val_a.lower() or "nilai total" in val_b.lower():
            if nilai_total.get("apip") is not None:
                write_cell(ws, row_idx, TOTAL_COL["nilai_total_apip"],
                           nilai_total["apip"])
            continue

        # --- Komponen (kode format "1.0", "2.0", dst. di kolom A) ---
        if re.match(r'^[1-4]\.0$', val_a) and val_a in index_komp:
            current_subkomp_kode = None  # Reset context
            komp = index_komp[val_a]
            penilaian = komp.get("penilaian_apip", {})
            write_cell(ws, row_idx, KOMP_COL["nilai_apip"],
                       penilaian.get("nilai"))
            write_cell(ws, row_idx, KOMP_COL["persen_apip"],
                       penilaian.get("persen"))
            write_cell(ws, row_idx, KOMP_COL["catatan_evaluator"],
                       komp.get("catatan_evaluator"))
            write_cell(ws, row_idx, KOMP_COL["rekomendasi_evaluator"],
                       komp.get("rekomendasi_evaluator"))
            continue

        # --- Sub-komponen (kode format "1.a", "1.b", dst. di kolom B) ---
        if re.match(r'^[1-4]\.[a-z]$', val_b) and not val_a and val_b in index_subkomp:
            current_subkomp_kode = val_b  # Update context
            sub = index_subkomp[val_b]
            penilaian = sub.get("penilaian_apip", {})
            write_cell(ws, row_idx, SUBKOMP_COL["persen_terpenuhi"],
                       penilaian.get("persen_terpenuhi"))
            write_cell(ws, row_idx, SUBKOMP_COL["nilai_apip"],
                       penilaian.get("nilai"))
            write_cell(ws, row_idx, SUBKOMP_COL["predikat_apip"],
                       penilaian.get("predikat"))
            write_cell(ws, row_idx, SUBKOMP_COL["nilai_akhir_apip"],
                       penilaian.get("nilai_akhir"))
            write_cell(ws, row_idx, SUBKOMP_COL["persen_apip"],
                       penilaian.get("persen_terpenuhi"))
            continue

        # --- Kriteria (nomor format "1)", "2)", dst. di kolom C) ---
        # Gunakan konteks sub-komponen saat ini untuk pencocokan tepat
        if re.match(r'^\d+\)$', val_c) and not val_a and not re.match(r'^[1-4]\.[a-z]$', val_b):
            if current_subkomp_kode is not None:
                key = (current_subkomp_kode, val_c)
                krit = index_kriteria.get(key)
                if krit:
                    penilaian = krit.get("penilaian_apip_baru", {})
                    write_cell(ws, row_idx, KRITERIA_COL["predikat_apip"],
                               penilaian.get("predikat"))
                    write_cell(ws, row_idx, KRITERIA_COL["nilai_apip"],
                               penilaian.get("nilai"))
                    write_cell(ws, row_idx, KRITERIA_COL["keterangan_apip"],
                               penilaian.get("keterangan"))
            continue


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main():
    if len(sys.argv) < 4:
        print(
            "Penggunaan: python fill_lke_apip.py "
            "<lke_asli.xls> <evaluasi.json> <output.xlsx>"
        )
        sys.exit(1)

    xls_path = sys.argv[1]
    json_path = sys.argv[2]
    output_path = sys.argv[3]

    print(f"Membaca LKE asli: {xls_path}")
    wb_xlsx, wb_xls = copy_xls_to_xlsx(xls_path)

    print(f"Membaca JSON evaluasi: {json_path}")
    with open(json_path, "r", encoding="utf-8") as f:
        data_json = json.load(f)

    # Isi sheet pertama (sheet LKE utama)
    ws = wb_xlsx.active
    if ws is None:
        ws = wb_xlsx.worksheets[0]

    print("Mengisi kolom evaluasi APIP...")
    isi_lke_dari_json(ws, data_json)

    print(f"Menyimpan hasil ke: {output_path}")
    wb_xlsx.save(output_path)
    print("Selesai.")


if __name__ == "__main__":
    main()
