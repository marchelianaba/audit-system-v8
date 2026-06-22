#!/usr/bin/env python3
"""
read_local_bukti.py — Ekstrak teks dari bukti dukung lokal, augment JSON LKE

Alur:
  1. Baca lke_extracted.json (output extract_lke.py)
  2. Untuk setiap kriteria, scan folder bukti_dukung/{sk_kode}/kr{nomor}/
     Jika tidak ada subfolder kriteria, baca semua file langsung dari {sk_kode}/
  3. Ekstrak teks dari PDF, data dari xlsx, daftar isi dari zip
  4. Simpan ke lke_with_bukti.json (siap dibaca Claude untuk evaluasi)

Struktur folder yang didukung:
  OPSI A — per kriteria:    bukti_dukung/1_a/kr1/file.pdf
  OPSI B — flat per unsur:  bukti_dukung/1_a/file.pdf  (semua kriteria 1.a baca folder yg sama)

Penggunaan:
  python read_local_bukti.py <lke_extracted.json> [bukti_folder] [output.json]

Butuh: pip install pdfminer.six openpyxl
"""
import sys
import json
import os
import re
import zipfile

# ── PDF ──────────────────────────────────────────────────────────────────────
def extract_pdf_text(path, max_chars=3000):
    """Ekstrak teks dari PDF. Potong di max_chars agar tidak membebani konteks."""
    try:
        from pdfminer.high_level import extract_text
        text = extract_text(path)
        if not text:
            return "[PDF tidak mengandung teks yang dapat diekstrak — kemungkinan scan/gambar]"
        # Bersihkan whitespace berlebihan
        text = re.sub(r'\n{3,}', '\n\n', text.strip())
        text = re.sub(r'[ \t]{2,}', ' ', text)
        if len(text) > max_chars:
            text = text[:max_chars] + f"\n... [terpotong, total {len(text)} karakter]"
        return text
    except Exception as e:
        return f"[Gagal ekstrak PDF: {e}]"

# ── XLSX ─────────────────────────────────────────────────────────────────────
def extract_xlsx_summary(path, max_rows=30):
    """Ekstrak ringkasan dari Excel: nama sheet + beberapa baris pertama."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        lines = []
        for sh_name in wb.sheetnames[:3]:  # maks 3 sheet
            ws = wb[sh_name]
            lines.append(f"[Sheet: {sh_name}]")
            count = 0
            for row in ws.iter_rows(values_only=True):
                cells = [str(c).strip() if c is not None else '' for c in row]
                non_empty = [c for c in cells if c]
                if non_empty:
                    lines.append(' | '.join(non_empty[:8]))
                    count += 1
                if count >= max_rows:
                    lines.append(f"... (baris selanjutnya tidak ditampilkan)")
                    break
        wb.close()
        return '\n'.join(lines)
    except Exception as e:
        return f"[Gagal baca xlsx: {e}]"

# ── ZIP ──────────────────────────────────────────────────────────────────────
def extract_zip_listing(path):
    """Daftar isi ZIP."""
    try:
        with zipfile.ZipFile(path, 'r') as z:
            names = z.namelist()
            info_lines = [f"[ZIP berisi {len(names)} file:]"]
            for name in names[:20]:
                size = z.getinfo(name).file_size
                info_lines.append(f"  - {name} ({size//1024} KB)")
            if len(names) > 20:
                info_lines.append(f"  ... dan {len(names)-20} file lainnya")
            return '\n'.join(info_lines)
    except Exception as e:
        return f"[Gagal baca zip: {e}]"

# ── DISPATCH ─────────────────────────────────────────────────────────────────
def extract_file(path):
    """Ekstrak konten berdasarkan ekstensi file."""
    ext = os.path.splitext(path)[1].lower()
    size_kb = os.path.getsize(path) / 1024
    header = f"[{os.path.basename(path)} — {size_kb:.0f} KB]\n"

    if ext == '.pdf':
        return header + extract_pdf_text(path)
    elif ext in ('.xlsx', '.xls'):
        return header + extract_xlsx_summary(path)
    elif ext == '.zip':
        return header + extract_zip_listing(path)
    elif ext in ('.jpg', '.jpeg', '.png', '.gif'):
        return header + "[File gambar — konten tidak dapat diekstrak sebagai teks]"
    elif ext in ('.docx', '.doc'):
        return header + "[File Word — gunakan pandoc untuk ekstrak teks jika diperlukan]"
    else:
        return header + f"[Tipe file {ext} tidak didukung untuk ekstraksi otomatis]"

# ── FOLDER MAPPING ───────────────────────────────────────────────────────────
def find_sk_folder(bukti_base, sk_kode):
    """
    Cari folder sub-komponen dengan toleransi variasi nama.
    Contoh: '1.a' bisa jadi folder '1.a', '1_a', '1-a', '1a'.
    """
    variants = [
        sk_kode,
        sk_kode.replace('.', '_'),
        sk_kode.replace('.', '-'),
        sk_kode.replace('.', ''),
    ]
    for v in variants:
        path = os.path.join(bukti_base, v)
        if os.path.isdir(path):
            return path
    return None

def find_kr_folder(bukti_base, sk_kode, nomor):
    """
    Cari folder kriteria dengan toleransi variasi nama.

    Strategi (urutan prioritas):
    1. Cari subfolder kriteria: {sk}/kr{n}, {sk}/kriteria{n}, {sk}/{n}
    2. Jika tidak ada → fallback ke folder sk langsung (mode flat)
       Dalam mode flat, semua kriteria dalam satu sub-komponen berbagi file yang sama.

    Returns (folder_path, is_flat):
      folder_path : str path ke folder yang berisi file, atau None jika tidak ada
      is_flat     : True jika menggunakan folder sk langsung (bukan subfolder kr)
    """
    n = nomor.rstrip(')').strip()

    # Variasi nama folder sub-komponen
    sk_variants = [
        sk_kode,
        sk_kode.replace('.', '_'),
        sk_kode.replace('.', '-'),
        sk_kode.replace('.', ''),
    ]
    # Variasi nama folder kriteria
    kr_variants = [f"kr{n}", f"kriteria{n}", f"{n}"]

    # Prioritas 1: cari subfolder kr
    for sk_v in sk_variants:
        for kr_v in kr_variants:
            path = os.path.join(bukti_base, sk_v, kr_v)
            if os.path.isdir(path):
                return path, False

    # Prioritas 2: fallback ke folder sk (mode flat)
    sk_path = find_sk_folder(bukti_base, sk_kode)
    if sk_path and any(os.path.isfile(os.path.join(sk_path, f)) for f in os.listdir(sk_path)):
        return sk_path, True

    return None, False

# ── MAIN ─────────────────────────────────────────────────────────────────────
def augment_lke(json_path, bukti_folder=None, output_path=None):
    with open(json_path, encoding='utf-8') as f:
        data = json.load(f)

    # Tentukan folder bukti_dukung
    if bukti_folder is None:
        bukti_folder = os.path.join(os.path.dirname(json_path), 'bukti_dukung')

    if not os.path.isdir(bukti_folder):
        print(f"ERROR: Folder bukti dukung tidak ditemukan: {bukti_folder}")
        print("Pastikan folder bukti_dukung/ sudah disiapkan di dalam folder penugasan.")
        sys.exit(1)

    if output_path is None:
        output_path = json_path.replace('.json', '_with_bukti.json')

    total_kr = 0
    kr_with_bukti = 0
    total_files = 0

    # Cache file flat per sk agar tidak re-extract untuk setiap kriteria
    flat_cache = {}  # sk_kode → list of {nama_file, path_lokal, konten}

    for k in data['komponen']:
        for sk in k['sub_komponen']:
            sk_kode = sk['kode']
            for kr in sk['kriteria']:
                total_kr += 1
                nomor = kr['nomor']
                kr_folder, is_flat = find_kr_folder(bukti_folder, sk_kode, nomor)

                if kr_folder is None:
                    kr['analisis_dokumen'] = []
                    continue

                if is_flat:
                    # Mode flat: semua kriteria dalam sk berbagi file yang sama
                    if sk_kode not in flat_cache:
                        flat_cache[sk_kode] = []
                        files = sorted(os.listdir(kr_folder))
                        for fname in files:
                            fpath = os.path.join(kr_folder, fname)
                            if not os.path.isfile(fpath):
                                continue
                            total_files += 1
                            teks = extract_file(fpath)
                            flat_cache[sk_kode].append({
                                "nama_file": fname,
                                "path_lokal": fpath,
                                "konten": teks
                            })
                            print(f"  FLAT {sk_kode}/ → {fname[:50]}")
                    dokumen = flat_cache[sk_kode]
                    if dokumen:
                        kr['analisis_dokumen'] = dokumen
                        kr['mode_folder'] = 'flat'
                        kr_with_bukti += 1
                    else:
                        kr['analisis_dokumen'] = []
                else:
                    # Mode per-kriteria: setiap kriteria punya subfolder sendiri
                    files = sorted(os.listdir(kr_folder))
                    dokumen = []
                    for fname in files:
                        fpath = os.path.join(kr_folder, fname)
                        if not os.path.isfile(fpath):
                            continue
                        total_files += 1
                        teks = extract_file(fpath)
                        dokumen.append({
                            "nama_file": fname,
                            "path_lokal": fpath,
                            "konten": teks
                        })
                        print(f"  OK  {sk_kode}/kr{nomor.rstrip(')')} → {fname[:50]}")
                    kr['analisis_dokumen'] = dokumen
                    if dokumen:
                        kr_with_bukti += 1

    # Simpan JSON augmented
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"\n=== SELESAI ===")
    print(f"Total kriteria    : {total_kr}")
    print(f"Kriteria ada bukti: {kr_with_bukti}")
    print(f"Total file dibaca : {total_files}")
    print(f"Output            : {output_path}")
    return output_path

def main():
    if len(sys.argv) < 2:
        print("Penggunaan: python read_local_bukti.py <lke_extracted.json> [bukti_folder] [output.json]")
        sys.exit(1)

    json_path  = sys.argv[1]
    bukti_dir  = sys.argv[2] if len(sys.argv) > 2 else None
    output     = sys.argv[3] if len(sys.argv) > 3 else None

    augment_lke(json_path, bukti_dir, output)

if __name__ == '__main__':
    main()
